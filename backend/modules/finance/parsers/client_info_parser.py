"""
Parser para ficheiro InfoClientes (infocliente.xlsx)
Enriquecimento semanal de dados de clientes.

Estrutura: Tabela plana, 1 linha = 1 cliente
"""
import logging
from typing import List, Dict, Any, Optional
from io import BytesIO
import openpyxl
from .account_normalizer import normalize_account_to_client_code

logger = logging.getLogger(__name__)


def parse_float(value: Any) -> float:
    """Converte valor para float"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(',', '.').replace(' ', '')
        try:
            return float(cleaned)
        except:
            return 0.0
    return 0.0


def parse_client_info(file_content: bytes) -> Dict[str, Any]:
    """
    Parse do ficheiro InfoClientes.
    
    Campos esperados:
    - CodCliente, Alm., Conta, Cliente, Saldo Conta, Carteira, Fat. Ano,
    - Domiciliações, Risco, Risco Seg., % Risco Seg., Albaranado
    
    Returns:
        Dict com:
        - clients: Lista de clientes
        - totals: Totais agregados
        - warnings: Avisos
        - errors: Erros
    """
    result = {
        'clients': [],
        'totals': {
            'client_count': 0,
            'rows_processed': 0,
            'total_balance': 0.0,
            'total_annual_revenue': 0.0,
            'total_pending_delivery': 0.0,
        },
        'warnings': [],
        'errors': []
    }
    
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        
        # Encontrar cabeçalho
        header_row = None
        header_mapping = {}
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_list = list(row) if row else []
            row_str = [str(c).strip() if c else '' for c in row_list]
            
            # Verificar se é linha de cabeçalho
            if 'CodCliente' in row_str or 'Cliente' in row_str:
                header_row = row_num
                for idx, col_name in enumerate(row_str):
                    if col_name:
                        # Normalizar nomes de colunas
                        normalized = col_name.replace('.', '').strip()
                        header_mapping[normalized] = idx
                        header_mapping[col_name] = idx  # Manter original também
                break
        
        if not header_row:
            result['errors'].append("Cabeçalho não encontrado no ficheiro")
            return result
        
        # Processar dados
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            row_list = list(row) if row else []
            
            # Skip linhas vazias
            if not any(c for c in row_list if c is not None and str(c).strip()):
                continue
            
            result['totals']['rows_processed'] += 1
            
            # Extrair valores usando mapeamento de colunas
            def get_val(col_names: List[str]) -> Any:
                for col_name in col_names:
                    idx = header_mapping.get(col_name)
                    if idx is not None and idx < len(row_list):
                        val = row_list[idx]
                        if val is not None:
                            return val
                return None
            
            # Iter 51: NUNCA usar CodPersona/Conta inteira. Extrair código
            # do cliente do sufixo da Conta (21111NNN → NNN).
            raw_code = str(get_val(['CodCliente', 'Cod Cliente', 'CodPersona']) or '').strip()
            account_raw = str(get_val(['Conta']) or '').strip()
            genes_code = normalize_account_to_client_code(account_raw)
            # Se não temos Conta reconhecível mas temos CodCliente explícito,
            # aceitamos esse (fluxo alternativo para ficheiros GENES antigos).
            if not genes_code and raw_code and not raw_code.startswith('21111'):
                genes_code = raw_code.lstrip('0') or raw_code
            if not genes_code:
                if account_raw:
                    result['warnings'].append(
                        f"Conta não reconhecida (padrão 21111NNN esperado): {account_raw!r}"
                    )
                continue
            
            # Skip linha de totais (última linha com soma)
            name = str(get_val(['Cliente']) or '').strip()
            if name.lower() in ['total', 'soma', 'totais', '']:
                continue
            
            balance = parse_float(get_val(['Saldo Conta', 'Saldo']))
            annual_revenue = parse_float(get_val(['Fat. Ano', 'Fat Ano', 'Faturação Ano']))
            pending_delivery = parse_float(get_val(['Albaranado']))
            
            # Risco: guarda sempre o valor raw. Detecta placeholder (>1M€) —
            # ficheiro GENES por vezes traz limites 999.999.999€ = "sem limite
            # atribuído", que não deve entrar no semáforo automaticamente.
            risk_raw = parse_float(get_val(['Risco']))
            risk_placeholder = abs(risk_raw) > 1_000_000
            risk_validated = 0.0 if risk_placeholder else risk_raw
            
            client = {
                'genes_code': genes_code,
                'warehouse': str(get_val(['Alm.', 'Alm', 'Armazém']) or '').strip(),
                'account': str(get_val(['Conta']) or '').strip(),
                'name': name,
                'saldo_conta': balance,
                'saldo_efec': parse_float(get_val(['Saldo Efec.', 'Saldo Efec'])),
                'saldo_desc': parse_float(get_val(['Saldo Desc.', 'Saldo Desc'])),
                'saldo_dev': parse_float(get_val(['Saldo Dev.', 'Saldo Dev'])),
                'portfolio': parse_float(get_val(['Carteira'])),
                'domiciliations': parse_float(get_val(['Domiciliações', 'Domiciliacoes'])),
                'risk_raw': risk_raw,
                'risk_validated': risk_validated,
                'risk_placeholder': risk_placeholder,
                'insured_risk_value': parse_float(get_val(['Risco Seg.', 'Risco Seg'])),
                'risk_percentage': parse_float(get_val(['% Risco Seg.', '% Risco Seg', 'Risco %'])),
                'pending_delivery': pending_delivery,
                'payment_method': str(get_val(['Forma Pagamento', 'Forma de Pagamento']) or '').strip(),
                'events_raw': str(get_val(['Eventos']) or '').strip(),
                # Legacy fields kept para retro-compatibilidade
                'total_balance': balance,
                'annual_revenue': annual_revenue,
                'risk_value': risk_validated,
            }
            
            result['clients'].append(client)
            result['totals']['client_count'] += 1
            result['totals']['total_balance'] += balance
            result['totals']['total_annual_revenue'] += annual_revenue
            result['totals']['total_pending_delivery'] += pending_delivery
        
        wb.close()
        
        logger.info(f"Parsed client info: {result['totals']['client_count']} clients")
        
    except Exception as e:
        logger.error(f"Error parsing client info: {e}")
        result['errors'].append(f"Erro ao processar ficheiro: {str(e)}")
    
    return result
