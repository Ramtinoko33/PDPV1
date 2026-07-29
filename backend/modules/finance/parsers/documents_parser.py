"""
Parser para ficheiro de Mapa de Documentos em Aberto (mapa.xlsx)
Mapa diário de documentos em aberto.

Estrutura: Tabela plana, 1 linha = 1 documento
"""
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, date
from io import BytesIO
import openpyxl
import re
from .account_normalizer import normalize_account_to_client_code

logger = logging.getLogger(__name__)

# Colunas esperadas
EXPECTED_COLUMNS = [
    'CodPersona', 'Conta', 'Tipo D. Pagamento', 'Forma Pagamento',
    'Data Fat.', 'Data Venc.', 'Cliente', 'Descritivo', 'Saldo',
    'Quantia', 'Vencido', 'Cobrado'
]


def parse_date(value: Any) -> Optional[str]:
    """Converte valor para data ISO string"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.split('.')[0].replace(' ', 'T'))
            return dt.date().isoformat()
        except:
            pass
        try:
            return datetime.strptime(value[:10], '%Y-%m-%d').date().isoformat()
        except:
            pass
    return None


def parse_float(value: Any) -> float:
    """Converte valor para float"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(',', '.').replace(' ', '')
        try:
            return float(cleaned)
        except:
            return 0.0
    return 0.0


def extract_document_info(descritivo: str) -> Dict[str, str]:
    """
    Extrai tipo e número do documento do campo Descritivo.
    Exemplo: "VTO. FAT./FT 026/3119" -> {"type": "FT", "number": "026/3119"}
    Exemplo: "VTO. FAT./NC C024/133" -> {"type": "NC", "number": "C024/133"}
    """
    result = {'type': 'UNKNOWN', 'number': ''}
    
    if not descritivo:
        return result
    
    descritivo = str(descritivo).upper()
    
    # Padrão: VTO. FAT./XX NUMERO
    # Onde XX pode ser FT, NC, etc.
    patterns = [
        r'VTO\.\s*FAT\./(\w+)\s+(.+)',  # VTO. FAT./FT 026/3119
        r'/(\w+)\s+(.+)',               # /FT 026/3119
        r'(\w+)\s+(\d+/\d+)',           # FT 026/3119
    ]
    
    for pattern in patterns:
        match = re.search(pattern, descritivo)
        if match:
            doc_type = match.group(1).strip()
            doc_number = match.group(2).strip()
            
            # Normalizar tipo
            if doc_type in ['FT', 'FAT', 'FATURA']:
                doc_type = 'FT'
            elif doc_type in ['NC', 'NOTA', 'CREDITO']:
                doc_type = 'NC'
            elif doc_type in ['RC', 'RECIBO']:
                doc_type = 'RC'
            
            result['type'] = doc_type
            result['number'] = doc_number
            break
    
    return result


def parse_open_documents(file_content: bytes) -> Dict[str, Any]:
    """
    Parse do ficheiro de mapa de documentos em aberto.
    
    Returns:
        Dict com:
        - documents: Lista de documentos
        - clients: Resumo por cliente
        - totals: Totais agregados
        - warnings: Avisos
        - errors: Erros
    """
    result = {
        'documents': [],
        'clients': {},  # Dict por genes_code
        'totals': {
            'document_count': 0,
            'client_count': 0,
            'total_balance': 0.0,
            'total_amount': 0.0,
            'total_overdue': 0.0,
            'credit_notes_count': 0,
            'credit_notes_amount': 0.0,
        },
        'warnings': [],
        'errors': []
    }
    
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        
        # Encontrar cabeçalho
        header_row = None
        header_mapping = {}
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_list = list(row) if row else []
            row_str = [str(c).strip() if c else '' for c in row_list]
            
            # Verificar se é linha de cabeçalho
            if 'CodPersona' in row_str or 'Cliente' in row_str:
                header_row = row_num
                for idx, col_name in enumerate(row_str):
                    if col_name:
                        header_mapping[col_name] = idx
                break
        
        if not header_row:
            result['errors'].append("Cabeçalho não encontrado no ficheiro")
            return result
        
        # Processar dados
        processed_clients = set()
        unmapped_accounts = set()
        
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            row_list = list(row) if row else []
            
            # Skip linhas vazias
            if not any(c for c in row_list if c is not None and str(c).strip()):
                continue
            
            # Extrair valores usando mapeamento de colunas
            def get_val(col_name: str) -> Any:
                idx = header_mapping.get(col_name)
                if idx is not None and idx < len(row_list):
                    return row_list[idx]
                return None
            
            # Iter 51: NUNCA usar CodPersona como client_key — extrair sempre
            # de Conta (conta contabilística) via normalizador.
            account_raw = str(get_val('Conta') or '').strip()
            genes_code = normalize_account_to_client_code(account_raw)
            if not genes_code:
                if account_raw:
                    unmapped_accounts.add(account_raw)
                continue
            
            # Dados do documento
            descritivo = str(get_val('Descritivo') or '')
            doc_info = extract_document_info(descritivo)
            
            amount = parse_float(get_val('Quantia'))
            overdue = parse_float(get_val('Vencido'))
            balance = parse_float(get_val('Saldo'))
            
            # Detectar nota de crédito (valor negativo)
            is_credit_note = amount < 0 or doc_info['type'] == 'NC'
            if is_credit_note:
                doc_info['type'] = 'NC'
            
            document = {
                'genes_code': genes_code,
                'account': account_raw,
                'client_name': str(get_val('Cliente') or '').strip(),
                'payment_type': str(get_val('Tipo D. Pagamento') or '').strip(),
                'payment_terms': str(get_val('Forma Pagamento') or '').strip(),
                'invoice_date': parse_date(get_val('Data Fat.')),
                'due_date': parse_date(get_val('Data Venc.')),
                'description': descritivo,
                'document_type': doc_info['type'],
                'document_number': doc_info['number'] or descritivo,
                'client_balance': balance,
                'amount': amount,
                'amount_overdue': overdue,
                'amount_collected': parse_float(get_val('Cobrado')),
                'is_credit_note': is_credit_note,
            }
            
            # Calcular dias vencidos
            if document['due_date']:
                try:
                    due = datetime.fromisoformat(document['due_date']).date()
                    today = date.today()
                    days = (today - due).days
                    document['days_overdue'] = max(0, days)
                except:
                    document['days_overdue'] = 0
            else:
                document['days_overdue'] = 0
            
            result['documents'].append(document)
            result['totals']['document_count'] += 1
            result['totals']['total_amount'] += amount
            result['totals']['total_overdue'] += overdue
            
            if is_credit_note:
                result['totals']['credit_notes_count'] += 1
                result['totals']['credit_notes_amount'] += abs(amount)
            
            # Agregar por cliente
            if genes_code not in result['clients']:
                result['clients'][genes_code] = {
                    'genes_code': genes_code,
                    'account': document['account'],
                    'name': document['client_name'],
                    'payment_terms': document['payment_terms'],
                    'total_balance': balance,  # Saldo do cliente (repetido por linha)
                    'document_count': 0,
                    'total_amount': 0.0,
                    'total_overdue': 0.0,
                    'credit_amount': 0.0,
                }
                processed_clients.add(genes_code)
            
            client = result['clients'][genes_code]
            client['document_count'] += 1
            client['total_amount'] += amount
            client['total_overdue'] += overdue
            if is_credit_note:
                client['credit_amount'] += abs(amount)
        
        result['totals']['client_count'] = len(processed_clients)
        
        if unmapped_accounts:
            sample = list(sorted(unmapped_accounts))[:5]
            result['warnings'].append(
                f"{len(unmapped_accounts)} conta(s) não correspondem ao padrão "
                f"21111NNN (linhas ignoradas). Exemplos: {sample}"
            )
        
        # Usar o maior saldo como total_balance (já que se repete por cliente)
        unique_balances = set()
        for doc in result['documents']:
            if doc['client_balance'] > 0:
                unique_balances.add((doc['genes_code'], doc['client_balance']))
        result['totals']['total_balance'] = sum(b for _, b in unique_balances)
        
        wb.close()
        
        logger.info(f"Parsed open documents: {result['totals']['client_count']} clients, {result['totals']['document_count']} documents")
        
    except Exception as e:
        logger.error(f"Error parsing open documents: {e}")
        result['errors'].append(f"Erro ao processar ficheiro: {str(e)}")
    
    return result
