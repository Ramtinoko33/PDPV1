"""
Parser para ficheiro de Evolução de Crédito Trimestral (evolucao.xlsx)
Análise de tendência trimestral (Fase futura).

Estrutura: Tabela pivot, colunas = trimestres (MM-YYYY)
"""
import logging
from typing import List, Dict, Any, Optional
from io import BytesIO
import openpyxl
import re

logger = logging.getLogger(__name__)


def parse_float(value: Any) -> float:
    """Converte valor para float"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(',', '.').replace(' ', '')
        try:
            return float(cleaned)
        except:
            return 0.0
    return 0.0


def parse_credit_evolution(file_content: bytes) -> Dict[str, Any]:
    """
    Parse do ficheiro de evolução de crédito trimestral.
    
    Colunas esperadas:
    - CODCLIENTE, Conta, Cliente
    - MM-YYYY (várias colunas com trimestres)
    
    Returns:
        Dict com:
        - clients: Lista de clientes com evolução
        - periods: Lista de períodos encontrados
        - totals: Totais agregados
        - warnings: Avisos
        - errors: Erros
    """
    result = {
        'clients': [],
        'periods': [],
        'totals': {
            'client_count': 0,
        },
        'warnings': [],
        'errors': []
    }
    
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        
        # Encontrar cabeçalho
        header_row = None
        header_mapping = {}
        period_columns = []  # Lista de (idx, period_str)
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_list = list(row) if row else []
            row_str = [str(c).strip() if c else '' for c in row_list]
            
            # Verificar se é linha de cabeçalho
            if 'CODCLIENTE' in row_str or 'CodCliente' in row_str or 'Cliente' in row_str:
                header_row = row_num
                
                for idx, col_name in enumerate(row_str):
                    if not col_name:
                        continue
                    
                    # Verificar se é coluna de período (MM-YYYY)
                    if re.match(r'^\d{2}-\d{4}$', col_name):
                        period_columns.append((idx, col_name))
                        result['periods'].append(col_name)
                    else:
                        # Normalizar nome
                        normalized = col_name.upper().replace('.', '').strip()
                        header_mapping[normalized] = idx
                        header_mapping[col_name] = idx
                break
        
        if not header_row:
            result['errors'].append("Cabeçalho não encontrado no ficheiro")
            return result
        
        if not period_columns:
            result['warnings'].append("Nenhuma coluna de período (MM-YYYY) encontrada")
        
        # Processar dados
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            row_list = list(row) if row else []
            
            # Skip linhas vazias
            if not any(c for c in row_list if c is not None and str(c).strip()):
                continue
            
            # Extrair valores usando mapeamento de colunas
            def get_val(col_names: List[str]) -> Any:
                for col_name in col_names:
                    idx = header_mapping.get(col_name)
                    if idx is not None and idx < len(row_list):
                        val = row_list[idx]
                        if val is not None:
                            return val
                return None
            
            genes_code = str(get_val(['CODCLIENTE', 'CodCliente', 'Cod Cliente']) or '').strip()
            if not genes_code:
                continue
            
            client = {
                'genes_code': genes_code,
                'account': str(get_val(['Conta', 'CONTA']) or '').strip(),
                'name': str(get_val(['Cliente', 'CLIENTE']) or '').strip(),
                'evolution': {}  # period -> balance
            }
            
            # Extrair valores de cada período
            for idx, period in period_columns:
                if idx < len(row_list):
                    balance = parse_float(row_list[idx])
                    client['evolution'][period] = balance
            
            # Calcular tendência (último vs primeiro)
            if len(client['evolution']) >= 2:
                sorted_periods = sorted(client['evolution'].keys())
                first_val = client['evolution'][sorted_periods[0]]
                last_val = client['evolution'][sorted_periods[-1]]
                
                if first_val > 0:
                    client['trend_percentage'] = ((last_val - first_val) / first_val) * 100
                else:
                    client['trend_percentage'] = 0.0
                
                client['trend_absolute'] = last_val - first_val
            
            result['clients'].append(client)
            result['totals']['client_count'] += 1
        
        wb.close()
        
        logger.info(f"Parsed credit evolution: {result['totals']['client_count']} clients, {len(result['periods'])} periods")
        
    except Exception as e:
        logger.error(f"Error parsing credit evolution: {e}")
        result['errors'].append(f"Erro ao processar ficheiro: {str(e)}")
    
    return result
