"""
Iteration 43 — Safety guards for open_documents import.

User-reported bug: importing a wrong/empty "Documentos Aberto" xlsx over an
existing dataset wiped finance_open_documents and generated hundreds of
`probable_payment` recovery events, inflating "Recuperado Hoje" to ~772k€.

This test file exercises the two backend safety guards + the same-day
re-import cleanup patch. All test data uses TST9x genes_codes and is
restored via snapshot/restore fixtures.
"""
import os
import io
import uuid
import asyncio
from datetime import datetime, timezone, date
from typing import List, Dict

import pytest
import requests
from openpyxl import Workbook

from dotenv import load_dotenv
load_dotenv('/app/backend/.env')

from motor.motor_asyncio import AsyncIOMotorClient  # noqa

# ---------------- env / auth ----------------

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
if not BASE_URL:
    with open('/app/frontend/.env') as f:
        for line in f:
            if line.startswith('REACT_APP_BACKEND_URL='):
                BASE_URL = line.split('=', 1)[1].strip().rstrip('/')

ADMIN_EMAIL = 'admin@pdpv.pt'
ADMIN_PASSWORD = os.environ.get('TEST_ADMIN_PASSWORD', 'HCNMEnKMLq')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']

TODAY = date.today().isoformat()

# Test genes codes reserved for this suite (must be unique to avoid collision
# with real data or with test_finance_phase2.TST90/TST91).
TST_G1 = '9993'
TST_G2 = '9994'
TST_G3 = '9995'
TST_GS_PREFIX = 'TSTB'   # for shrinkage bulk seed

session = requests.Session()
session.headers.update({'Content-Type': 'application/json'})


def _login():
    r = session.post(f'{BASE_URL}/api/auth/login',
                     json={'email': ADMIN_EMAIL, 'password': ADMIN_PASSWORD})
    assert r.status_code == 200, f'login: {r.status_code} {r.text}'
    session.headers.update({'Authorization': f'Bearer {r.json()["token"]}'})


@pytest.fixture(scope='session', autouse=True)
def _auth():
    _login()


# ---------------- helpers ----------------

async def _db():
    c = AsyncIOMotorClient(MONGO_URL)
    return c[DB_NAME]


def _build_open_docs_xlsx(rows: List[Dict], marker: str) -> bytes:
    """Build a xlsx compatible with parse_open_documents (mirror of existing test)."""
    wb = Workbook()
    ws = wb.active
    headers = ['CodPersona', 'Conta', 'Tipo D. Pagamento', 'Forma Pagamento',
               'Data Fat.', 'Data Venc.', 'Cliente', 'Descritivo', 'Saldo',
               'Registo B.', 'Registo C.', 'Registo D.', 'Quantia', 'Vencido',
               'Cobrado', 'Estado', 'Eventos']
    ws.append(headers)
    for r in rows:
        # Iter 51: Conta segue padrão 21111NNN (5 dígitos fixos + código).
        # Construímos a Conta a partir do genes_code do teste para que o
        # parser normalizador devolva o mesmo genes_code.
        gc = str(r['genes_code'])
        # gc pode ser numérico ou não — se for numérico usamos como sufixo,
        # senão inventamos um sufixo determinístico.
        try:
            int(gc)
            conta_sufixo = gc.zfill(5)
        except ValueError:
            conta_sufixo = str(abs(hash(gc)) % 100000).zfill(5)
        conta = r.get('conta', f'21111{conta_sufixo}')
        ws.append([
            r['genes_code'], conta, 'TR', '30D',
            r.get('data_fat', '2025-11-01'), r.get('data_venc', '2025-12-01'),
            r['client_name'], r['descritivo'], r['saldo'],
            '', '', '', r['quantia'], r['vencido'], r.get('cobrado', 0), 'Aberto', marker
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_empty_open_docs_xlsx(marker: str) -> bytes:
    """Header only, zero data rows — reproduces the reported bug scenario."""
    wb = Workbook()
    ws = wb.active
    headers = ['CodPersona', 'Conta', 'Tipo D. Pagamento', 'Forma Pagamento',
               'Data Fat.', 'Data Venc.', 'Cliente', 'Descritivo', 'Saldo',
               'Registo B.', 'Registo C.', 'Registo D.', 'Quantia', 'Vencido',
               'Cobrado', 'Estado', 'Eventos']
    ws.append(headers)
    # Ghost cell so file hash is unique per marker (avoids "already imported")
    ws['Z1'] = marker
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(content: bytes, filename: str):
    headers = {'Authorization': session.headers['Authorization']}
    files = {'file': (filename, content,
                      'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    return requests.post(
        f'{BASE_URL}/api/finance/imports/open_documents',
        files=files, headers=headers
    )


# ---------------- snapshot / restore ----------------

@pytest.fixture(scope='module', autouse=True)
def isolate_finance_state():
    """Snapshot finance_open_documents, finance_recovery_events, data_health.
    Restore at end + cleanup TST imports and their files."""
    async def snap():
        db = await _db()
        docs = await db.finance_open_documents.find({}, {'_id': 0}).to_list(20000)
        recovery = await db.finance_recovery_events.find({}, {'_id': 0}).to_list(20000)
        health = await db.finance_data_health.find_one(
            {'source_type': 'open_documents'}, {'_id': 0}
        )
        return {'docs': docs, 'recovery': recovery, 'health': health}
    state = asyncio.run(snap())

    yield state

    async def restore():
        db = await _db()
        # Restore open_documents to original
        await db.finance_open_documents.delete_many({})
        if state['docs']:
            await db.finance_open_documents.insert_many([dict(d) for d in state['docs']])
        # Restore recovery_events: delete anything from today (the tests
        # dirtied it) then re-insert original events that were on that date.
        await db.finance_recovery_events.delete_many({'date': TODAY})
        orig_today = [dict(e) for e in state['recovery'] if e.get('date') == TODAY]
        if orig_today:
            await db.finance_recovery_events.insert_many(orig_today)
        # Also purge any residual TST_ events on other dates
        await db.finance_recovery_events.delete_many(
            {'genes_code': {'$in': [TST_G1, TST_G2, TST_G3]}}
        )
        await db.finance_recovery_events.delete_many(
            {'genes_code': {'$regex': f'^{TST_GS_PREFIX}'}}
        )
        # Restore data_health
        if state['health']:
            await db.finance_data_health.update_one(
                {'source_type': 'open_documents'},
                {'$set': state['health']}, upsert=True
            )
        # Delete test import records + their files
        cur = db.finance_imports.find(
            {'filename': {'$regex': '^TSTGUARD-'}},
            {'id': 1, 'original_file_path': 1}
        )
        async for imp in cur:
            p = imp.get('original_file_path')
            if p and os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass
        await db.finance_imports.delete_many({'filename': {'$regex': '^TSTGUARD-'}})
    asyncio.run(restore())


def _seed_docs(rows: List[Dict]):
    """Manually put docs into finance_open_documents (bypasses import)."""
    async def _go():
        db = await _db()
        now = datetime.now(timezone.utc).isoformat()
        payload = []
        for r in rows:
            payload.append({
                'doc_key': f"{r['genes_code']}_{r['document_number']}",
                'genes_code': r['genes_code'],
                'client_name': r['client_name'],
                'document_number': r['document_number'],
                'document_type': 'FT',
                'description': r.get('description', f"VTO. FAT./FT {r['document_number']}"),
                'amount': r['amount'],
                'amount_overdue': r['amount'],
                'amount_collected': 0.0,
                'client_balance': r['amount'],
                'is_credit_note': False,
                'account': '21100001',
                'payment_type': 'TR',
                'payment_terms': '30D',
                'invoice_date': '2025-11-01',
                'due_date': '2025-12-01',
                'days_overdue': 30,
                'as_of_date': TODAY,
                'import_id': 'seed-' + uuid.uuid4().hex[:8],
                'updated_at': now,
            })
        await db.finance_open_documents.insert_many(payload)
    asyncio.run(_go())


def _clear_open_docs():
    async def _go():
        db = await _db()
        await db.finance_open_documents.delete_many({})
    asyncio.run(_go())


def _count_open_docs() -> int:
    async def _go():
        db = await _db()
        return await db.finance_open_documents.count_documents({})
    return asyncio.run(_go())


def _count_recovery_today() -> int:
    async def _go():
        db = await _db()
        return await db.finance_recovery_events.count_documents({'date': TODAY})
    return asyncio.run(_go())


def _sum_recovery_today() -> float:
    async def _go():
        db = await _db()
        total = 0.0
        async for e in db.finance_recovery_events.find({'date': TODAY}, {'_id': 0, 'amount': 1}):
            total += float(e.get('amount', 0) or 0)
        return total
    return asyncio.run(_go())


# =========================================================================
# Tests
# =========================================================================

class TestValidImport:
    """Import with >0 parsed docs on empty DB → imported, totals>0."""

    def test_valid_import_success(self):
        _clear_open_docs()
        marker = uuid.uuid4().hex[:8]
        rows = [
            {'genes_code': TST_G1, 'client_name': 'Cliente G1',
             'descritivo': f'VTO. FAT./FT 999/{marker}',
             'saldo': 100.0, 'quantia': 100.0, 'vencido': 100.0},
            {'genes_code': TST_G2, 'client_name': 'Cliente G2',
             'descritivo': f'VTO. FAT./FT 998/{marker}',
             'saldo': 200.0, 'quantia': 200.0, 'vencido': 200.0},
        ]
        content = _build_open_docs_xlsx(rows, marker)
        r = _upload(content, f'TSTGUARD-valid-{marker}.xlsx')
        assert r.status_code == 200, r.text
        data = r.json()
        assert data['success'] is True
        assert data['status'] in ('imported', 'accepted_with_warnings'), data
        assert data['totals']['clients'] == 2, data['totals']
        assert data['totals']['documents'] == 2, data['totals']


class TestGuard1CatastrophicEmpty:
    """When parser produces 0 docs and DB has docs → status=rejected,
    finance_open_documents unchanged, no new recovery events."""

    def test_empty_import_rejected_and_no_mutation(self):
        # Seed 5 docs
        _clear_open_docs()
        seed_rows = [
            {'genes_code': TST_G1, 'client_name': 'C1', 'document_number': f'999/{i}',
             'amount': 100.0 + i} for i in range(5)
        ]
        _seed_docs(seed_rows)
        assert _count_open_docs() == 5

        recovery_before = _count_recovery_today()

        # Upload an empty (header-only) xlsx — the reported bug scenario
        marker = uuid.uuid4().hex[:8]
        content = _build_empty_open_docs_xlsx(marker)
        r = _upload(content, f'TSTGUARD-empty-{marker}.xlsx')
        assert r.status_code == 200, r.text
        data = r.json()

        # (a) status = rejected
        assert data['status'] == 'rejected', data
        assert data['success'] is False, data
        # (b) errors contain a meaningful message
        errors_str = ' '.join(data.get('errors', []))
        assert len(errors_str) > 0
        assert '0 documentos' in errors_str or 'rejeitado' in errors_str.lower(), errors_str

        # (c) finance_open_documents NOT mutated — still 5 docs
        assert _count_open_docs() == 5, "Guard 1 must NOT delete existing docs"

        # (d) No new recovery events on today's date
        recovery_after = _count_recovery_today()
        assert recovery_after == recovery_before, (
            f"Guard 1 must NOT create recovery events (before={recovery_before}, "
            f"after={recovery_after})"
        )


class TestGuard2CatastrophicShrinkage:
    """Old count >= 100 AND new < 10% AND delta > 100 → pending_approval,
    NO mutation, NO recovery events."""

    def test_shrinkage_import_pending_approval_and_no_mutation(self):
        # Seed 150 test docs
        _clear_open_docs()
        seed_rows = [
            {'genes_code': f'{TST_GS_PREFIX}{i:03d}', 'client_name': f'Cli {i}',
             'document_number': f'AA/{i}', 'amount': 50.0 + i}
            for i in range(150)
        ]
        _seed_docs(seed_rows)
        assert _count_open_docs() == 150

        recovery_before = _count_recovery_today()

        # Upload a small xlsx with only 5 rows (< 10% of 150 = 15, delta=145 > 100)
        marker = uuid.uuid4().hex[:8]
        rows = [
            {'genes_code': f'{TST_GS_PREFIX}{i:03d}', 'client_name': f'Cli {i}',
             'descritivo': f'VTO. FAT./FT AA/{i}-{marker}',
             'saldo': 50.0 + i, 'quantia': 50.0 + i, 'vencido': 50.0 + i}
            for i in range(5)
        ]
        content = _build_open_docs_xlsx(rows, marker)
        r = _upload(content, f'TSTGUARD-shrink-{marker}.xlsx')
        assert r.status_code == 200, r.text
        data = r.json()

        assert data['status'] == 'pending_approval', data
        assert data['success'] is False, data
        warnings_str = ' '.join(data.get('warnings', []))
        assert 'edução' in warnings_str or 'aprovação' in warnings_str.lower(), warnings_str

        # No mutation
        assert _count_open_docs() == 150, "Guard 2 must NOT delete existing docs"
        # No recovery events
        assert _count_recovery_today() == recovery_before, (
            "Guard 2 must NOT create recovery events"
        )


class TestSameDayReimportCleanup:
    """Two open_documents imports the same day: 2nd wipes 1st's recovery
    events for that as_of and writes only its own."""

    def test_same_day_reimport_replaces_recovery_events(self):
        # Setup: seed 3 docs
        _clear_open_docs()
        seed = [
            {'genes_code': TST_G1, 'client_name': 'C1', 'document_number': f'A/1',
             'amount': 100.0},
            {'genes_code': TST_G2, 'client_name': 'C2', 'document_number': f'A/2',
             'amount': 200.0},
            {'genes_code': TST_G3, 'client_name': 'C3', 'document_number': f'A/3',
             'amount': 300.0},
        ]
        _seed_docs(seed)
        assert _count_open_docs() == 3

        # Clear any recovery events for today so we measure cleanly
        async def clear_today():
            db = await _db()
            await db.finance_recovery_events.delete_many({'date': TODAY})
        asyncio.run(clear_today())
        assert _count_recovery_today() == 0

        # First import: drops G2 & G3 → 2 probable_payment events (sum=500)
        marker1 = uuid.uuid4().hex[:8]
        rows1 = [{
            'genes_code': TST_G1, 'client_name': 'C1',
            'descritivo': 'VTO. FAT./FT A/1',
            'saldo': 100.0, 'quantia': 100.0, 'vencido': 100.0,
        }]
        content1 = _build_open_docs_xlsx(rows1, marker1)
        r1 = _upload(content1, f'TSTGUARD-reimp1-{marker1}.xlsx')
        assert r1.status_code == 200, r1.text
        data1 = r1.json()
        assert data1['success'] is True
        assert data1['totals']['probable_payments'] == 2, data1['totals']
        assert abs(data1['totals']['recovered_amount'] - 500.0) < 0.02

        assert _count_recovery_today() == 2
        sum_after_1 = _sum_recovery_today()
        assert abs(sum_after_1 - 500.0) < 0.02

        # Restore docs G2, G3 into finance_open_documents (simulate "dataset
        # would be intact if guard had blocked import1"; we still want to
        # verify cleanup on next same-day import).
        _seed_docs([
            {'genes_code': TST_G2, 'client_name': 'C2', 'document_number': f'A/2',
             'amount': 200.0},
            {'genes_code': TST_G3, 'client_name': 'C3', 'document_number': f'A/3',
             'amount': 300.0},
        ])
        assert _count_open_docs() == 3

        # Second import same day: again drops G2 & G3 → 2 events sum=500.
        # After the fix, delete_many({date:today}) must wipe the 2 previous
        # events before insert — final count remains 2, NOT 4.
        marker2 = uuid.uuid4().hex[:8]
        rows2 = [{
            'genes_code': TST_G1, 'client_name': 'C1',
            'descritivo': 'VTO. FAT./FT A/1',
            'saldo': 100.0, 'quantia': 100.0, 'vencido': 100.0,
        }]
        content2 = _build_open_docs_xlsx(rows2, marker2)
        r2 = _upload(content2, f'TSTGUARD-reimp2-{marker2}.xlsx')
        assert r2.status_code == 200, r2.text
        data2 = r2.json()
        assert data2['success'] is True

        final_count = _count_recovery_today()
        final_sum = _sum_recovery_today()
        assert final_count == 2, (
            f"Same-day re-import must replace events (got {final_count}, expected 2)"
        )
        assert abs(final_sum - 500.0) < 0.02, (
            f"Recovered sum must not accumulate (got {final_sum}, expected ~500)"
        )


class TestDashboardNotInflatedByRejection:
    """After a rejected empty import, dashboard's recovered_today must not
    reflect any inflated amount from the rejected import."""

    def test_dashboard_unchanged_after_rejected_empty(self):
        # Baseline dashboard
        r0 = session.get(f'{BASE_URL}/api/finance/dashboard')
        assert r0.status_code == 200
        base_recovered = r0.json().get('recovered_today', 0)

        # Seed 3 docs
        _clear_open_docs()
        _seed_docs([
            {'genes_code': TST_G1, 'client_name': 'C1', 'document_number': 'B/1', 'amount': 999.99},
            {'genes_code': TST_G2, 'client_name': 'C2', 'document_number': 'B/2', 'amount': 888.88},
            {'genes_code': TST_G3, 'client_name': 'C3', 'document_number': 'B/3', 'amount': 777.77},
        ])
        # Try empty import → must be rejected
        marker = uuid.uuid4().hex[:8]
        content = _build_empty_open_docs_xlsx(marker)
        r = _upload(content, f'TSTGUARD-dash-{marker}.xlsx')
        assert r.status_code == 200
        assert r.json()['status'] == 'rejected'

        # Dashboard must not have gained ~2666 (999.99+888.88+777.77) recovered
        r1 = session.get(f'{BASE_URL}/api/finance/dashboard')
        assert r1.status_code == 200
        new_recovered = r1.json().get('recovered_today', 0)
        # allow small deltas from parallel activity, but the huge sum must NOT be here
        assert new_recovered < base_recovered + 100, (
            f"Dashboard recovered_today inflated after rejected import "
            f"(before={base_recovered}, after={new_recovered})"
        )
