"""
Iteration 44 — Safety guards for overdue_balances (Saldos Vencidos) import.

User-reported bug (Feb 2026): a Saldos Vencidos xlsx uploaded to production
was parsed as 1492 clients / 0 documents (probably wrong file structure).
The pending approval, once clicked, threw HTTP 500 without any actionable
message and — if approved — would wipe every finance_documents record
because the OVERDUE flow had no catastrophic-empty guard (only Docs Aberto
had one, from iteration 43).

This test suite adds an equivalent guard for the OVERDUE_BALANCES flow:

- Guard 1: parsed docs == 0 AND existing DB has docs → status=rejected,
  finance_documents untouched. Guard runs even with force_approved=True
  so that approving a corrupt pending import cannot ever wipe data.

- Approve endpoint now returns HTTP 400 with the guard reason (previously
  it returned a generic HTTP 500).
"""
import os
import io
import uuid
import asyncio
from datetime import datetime, timezone, date
from typing import List, Dict

import pytest
import requests
from openpyxl import Workbook

from dotenv import load_dotenv
load_dotenv('/app/backend/.env')

from motor.motor_asyncio import AsyncIOMotorClient  # noqa

# ---------------- env / auth ----------------

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
if not BASE_URL:
    with open('/app/frontend/.env') as f:
        for line in f:
            if line.startswith('REACT_APP_BACKEND_URL='):
                BASE_URL = line.split('=', 1)[1].strip().rstrip('/')

ADMIN_EMAIL = 'admin@pdpv.pt'
ADMIN_PASSWORD = os.environ.get('TEST_ADMIN_PASSWORD', 'HCNMEnKMLq')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']

TODAY = date.today().isoformat()

# Test genes codes reserved for this suite
TST_G1 = 'TSTV1'
TST_G2 = 'TSTV2'

session = requests.Session()
session.headers.update({'Content-Type': 'application/json'})


def _login():
    r = session.post(f'{BASE_URL}/api/auth/login',
                     json={'email': ADMIN_EMAIL, 'password': ADMIN_PASSWORD})
    assert r.status_code == 200, f'login: {r.status_code} {r.text}'
    session.headers.update({'Authorization': f'Bearer {r.json()["token"]}'})


@pytest.fixture(scope='session', autouse=True)
def _auth():
    _login()


# ---------------- helpers ----------------

async def _db():
    c = AsyncIOMotorClient(MONGO_URL)
    return c[DB_NAME]


def _build_overdue_xlsx_with_docs(marker: str) -> bytes:
    """Build a well-formed Saldos Vencidos xlsx with 1 client + 1 doc."""
    wb = Workbook()
    ws = wb.active
    # Client header
    ws.append(['Cliente', 'Cód. Cliente', 'Localidade', 'Região', 'Email',
               'Telefone1', 'Telefone2', 'Importe Total Vencido', 'Saldo Cliente'])
    # Client row
    ws.append([f'Cliente Teste {marker}', TST_G1, 'Lisboa', 'Sul',
               'test@test.pt', '', '', 100.0, 100.0])
    # Doc header
    ws.append(['', 'Documento', 'Data da fatura', 'Data Vencimento', 'CódSede',
               'Sede', 'Dias Vencidos', 'Importe Vencimiento', 'Vencido Factura'])
    # Doc row
    ws.append(['', '026/9999', datetime(2025, 11, 1), datetime(2025, 12, 1),
               '01', 'Sede', 30, 100.0, 100.0])
    ws['Z1'] = marker
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_overdue_xlsx_zero_docs(marker: str) -> bytes:
    """Build a Saldos Vencidos xlsx where parser sees only client headers/rows
    but no document rows — reproduces the reported bug scenario (1492/0)."""
    wb = Workbook()
    ws = wb.active
    ws.append(['Cliente', 'Cód. Cliente', 'Localidade', 'Região', 'Email',
               'Telefone1', 'Telefone2', 'Importe Total Vencido', 'Saldo Cliente'])
    # A few "client" rows with no following document rows
    for i in range(5):
        ws.append([f'Cli {i} {marker}', f'ZERO{i:02d}', 'Lx', 'Sul',
                   None, None, None, 0.0, 0.0])
    ws['Z1'] = marker
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_overdue(content: bytes, filename: str):
    headers = {'Authorization': session.headers['Authorization']}
    files = {'file': (filename, content,
                      'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    return requests.post(
        f'{BASE_URL}/api/finance/imports/overdue_balances',
        files=files, headers=headers
    )


def _approve(import_id: str):
    return requests.post(
        f'{BASE_URL}/api/finance/imports/{import_id}/approve',
        headers={'Authorization': session.headers['Authorization']}
    )


# ---------------- snapshot / restore ----------------

@pytest.fixture(scope='module', autouse=True)
def isolate_finance_state():
    """Snapshot finance_documents, finance_clients, imports. Restore on teardown."""
    async def snap():
        db = await _db()
        docs = await db.finance_documents.find({}, {'_id': 0}).to_list(20000)
        clients = await db.finance_clients.find({}, {'_id': 0}).to_list(20000)
        return {'docs': docs, 'clients': clients}
    state = asyncio.run(snap())

    yield state

    async def restore():
        db = await _db()
        # Restore documents
        await db.finance_documents.delete_many({})
        if state['docs']:
            await db.finance_documents.insert_many([dict(d) for d in state['docs']])
        # Restore clients (only wipe test ones to be safe)
        await db.finance_clients.delete_many(
            {'genes_code': {'$in': [TST_G1, TST_G2] + [f'ZERO{i:02d}' for i in range(5)]}}
        )
        # Cleanup our test imports + their files
        cur = db.finance_imports.find(
            {'filename': {'$regex': '^TSTOVGUARD-'}},
            {'id': 1, 'original_file_path': 1}
        )
        async for imp in cur:
            p = imp.get('original_file_path')
            if p and os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass
        await db.finance_imports.delete_many({'filename': {'$regex': '^TSTOVGUARD-'}})
    asyncio.run(restore())


def _seed_docs(genes_codes: List[str]):
    """Insert finance_documents entries under given genes_codes (bypasses import)."""
    async def _go():
        db = await _db()
        now = datetime.now(timezone.utc).isoformat()
        payload = []
        for code in genes_codes:
            for i in range(3):
                payload.append({
                    'id': f'{code}_SEED{i:02d}',
                    'client_id': None,
                    'genes_code': code,
                    'document_type': 'FT',
                    'document_number': f'SEED{i:02d}',
                    'invoice_date': '2025-11-01',
                    'due_date': '2025-12-01',
                    'amount_original': 100.0,
                    'amount_open': 100.0,
                    'amount_overdue': 100.0,
                    'days_overdue': 30,
                    'classification': 'collectable',
                    'effective_classification': 'collectable',
                    'manually_marked_collectable': False,
                    'manual_action': None,
                    'last_import_id': 'seed-' + uuid.uuid4().hex[:8],
                    'created_at': now,
                    'updated_at': now,
                })
        if payload:
            await db.finance_documents.insert_many(payload)
    asyncio.run(_go())


def _clear_finance_docs():
    async def _go():
        db = await _db()
        await db.finance_documents.delete_many({})
    asyncio.run(_go())


def _count_finance_docs() -> int:
    async def _go():
        db = await _db()
        return await db.finance_documents.count_documents({})
    return asyncio.run(_go())


# =========================================================================
# Tests
# =========================================================================

class TestValidOverdueImport:
    """Well-formed overdue file with docs → imported, docs count > 0."""

    def test_valid_import_success(self):
        _clear_finance_docs()
        marker = uuid.uuid4().hex[:8]
        content = _build_overdue_xlsx_with_docs(marker)
        r = _upload_overdue(content, f'TSTOVGUARD-valid-{marker}.xlsx')
        assert r.status_code == 200, r.text
        data = r.json()
        # The safety guard MUST NOT trigger for a well-formed file with >0 docs.
        # The import may still land in pending_approval due to the diff threshold
        # against previous imports (that's not the guard). What we assert is:
        # (a) the response was not rejected by the guard.
        assert data['status'] != 'rejected', data
        # (b) the errors list does NOT contain the guard message.
        errors_str = ' '.join(data.get('errors', []))
        assert '0 documentos' not in errors_str, errors_str
        # (c) parsed documents count > 0.
        assert data['totals'].get('documents', 0) >= 1, data['totals']


class TestOverdueGuardCatastrophicEmpty:
    """Parser produces 0 documents while DB has docs → rejected, docs intact.

    This is the exact reported production scenario: 1492 clientes / 0 docs.
    """

    def test_zero_docs_import_rejected_and_no_mutation(self):
        # Seed 6 real documents (3 for each of 2 clients)
        _clear_finance_docs()
        _seed_docs([TST_G1, TST_G2])
        assert _count_finance_docs() == 6

        marker = uuid.uuid4().hex[:8]
        content = _build_overdue_xlsx_zero_docs(marker)
        r = _upload_overdue(content, f'TSTOVGUARD-zero-{marker}.xlsx')
        assert r.status_code == 200, r.text
        data = r.json()

        # (a) status = rejected + success=False
        assert data['status'] == 'rejected', data
        assert data.get('success') is False, data
        # (b) errors mention 0 documentos + saldos vencidos wording
        errors_str = ' '.join(data.get('errors', []))
        assert '0 documentos' in errors_str, errors_str
        assert 'rejeitado' in errors_str.lower() or 'reset' in errors_str.lower(), errors_str
        # (c) finance_documents NOT touched
        assert _count_finance_docs() == 6, "Guard MUST NOT wipe existing docs"


class TestOverdueApproveReturns400OnGuard:
    """Approving a corrupt PENDING import must be blocked by the guard and
    return HTTP 400 with a helpful detail (not a generic 500)."""

    def test_approve_pending_zero_docs_import_returns_400(self):
        _clear_finance_docs()
        _seed_docs([TST_G1])
        assert _count_finance_docs() == 3

        # Manually create a PENDING_APPROVAL import in DB, pointing to a
        # freshly-saved xlsx with 0 documents. This simulates the exact
        # state the production user is in.
        async def _prep():
            db = await _db()
            from pathlib import Path
            upload_dir = Path('/app/backend/uploads/finance')
            upload_dir.mkdir(parents=True, exist_ok=True)
            marker = uuid.uuid4().hex[:8]
            import_id = str(uuid.uuid4())
            content = _build_overdue_xlsx_zero_docs(marker)
            file_path = upload_dir / f'{import_id}.xlsx'
            file_path.write_bytes(content)
            await db.finance_imports.insert_one({
                'id': import_id,
                'type': 'overdue_balances',
                'source_method': 'manual_upload',
                'filename': f'TSTOVGUARD-pending-{marker}.xlsx',
                'file_hash': uuid.uuid4().hex,
                'as_of_date': TODAY,
                'uploaded_by': 'test',
                'uploaded_at': datetime.now(timezone.utc).isoformat(),
                'status': 'pending_approval',
                'original_file_path': str(file_path),
                'totals': {'clients': 5, 'documents': 0,
                           'total_balance': 0, 'total_overdue': 0},
                'warnings': ['manual seed'],
                'errors': [],
                'approved_by': None,
                'approved_at': None,
            })
            return import_id
        import_id = asyncio.run(_prep())

        docs_before = _count_finance_docs()

        r = _approve(import_id)
        # Should be 400 (guard blocks), not 500
        assert r.status_code == 400, (r.status_code, r.text)
        body = r.json()
        assert 'detail' in body, body
        assert '0 documentos' in body['detail'], body['detail']

        # Docs must be intact (guard prevented wipe)
        assert _count_finance_docs() == docs_before, "Approve MUST NOT wipe docs"
