"""
Iteration 46 — Import preview endpoint + safety guard preview.

Novo endpoint GET /api/finance/imports/{id}/preview que devolve os
números-chave antes do utilizador clicar em Aprovar:
  - clientes / documentos actuais vs novos
  - delta e diff_pct
  - guard_warnings (motivos pelos quais o approve seria bloqueado)
  - is_critical (bandeira para o frontend pintar aviso vermelho)

Este teste cobre:
  1. Preview de uma importação Saldos Vencidos VÁLIDA (>0 docs, sem guard warnings)
  2. Preview de uma importação Saldos Vencidos com 0 documentos → guard warning + is_critical=true
  3. Preview quando o import_id não existe → 404
"""
import os
import io
import uuid
import asyncio
from datetime import datetime, timezone
from pathlib import Path

import pytest
import requests
from openpyxl import Workbook

from dotenv import load_dotenv
load_dotenv('/app/backend/.env')

from motor.motor_asyncio import AsyncIOMotorClient  # noqa

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
if not BASE_URL:
    with open('/app/frontend/.env') as f:
        for line in f:
            if line.startswith('REACT_APP_BACKEND_URL='):
                BASE_URL = line.split('=', 1)[1].strip().rstrip('/')

ADMIN_EMAIL = 'admin@pdpv.pt'
ADMIN_PASSWORD = os.environ.get('TEST_ADMIN_PASSWORD', 'HCNMEnKMLq')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']

TST_G = 'TSTP1'

session = requests.Session()
session.headers.update({'Content-Type': 'application/json'})


def _login():
    r = session.post(f'{BASE_URL}/api/auth/login',
                     json={'email': ADMIN_EMAIL, 'password': ADMIN_PASSWORD})
    assert r.status_code == 200, r.text
    session.headers.update({'Authorization': f'Bearer {r.json()["token"]}'})


@pytest.fixture(scope='session', autouse=True)
def _auth():
    _login()


async def _db():
    c = AsyncIOMotorClient(MONGO_URL)
    return c[DB_NAME]


def _build_overdue_xlsx_valid(marker: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(['Cliente', 'Cód. Cliente', 'Localidade', 'Região', 'Email',
               'Telefone1', 'Telefone2', 'Importe Total Vencido', 'Saldo Cliente'])
    ws.append([f'Cli {marker}', TST_G, 'Lx', 'Sul', 'x@x.pt',
               None, None, 250.0, 250.0])
    ws.append(['', 'Documento', 'Data da fatura', 'Data Vencimento', 'CódSede',
               'Sede', 'Dias Vencidos', 'Importe Vencimiento', 'Vencido Factura'])
    ws.append(['', '999/prev1', datetime(2025, 11, 1), datetime(2025, 12, 1),
               '01', 'Sede', 30, 250.0, 250.0])
    ws['Z1'] = marker
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_overdue_xlsx_zero_docs(marker: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(['Cliente', 'Cód. Cliente', 'Localidade', 'Região', 'Email',
               'Telefone1', 'Telefone2', 'Importe Total Vencido', 'Saldo Cliente'])
    for i in range(3):
        ws.append([f'Cli{i}', f'ZERO{i}', 'Lx', 'Sul', None,
                   None, None, 0.0, 0.0])
    ws['Z1'] = marker
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _create_pending_import(content: bytes, filename: str, doc_count_hint: int) -> str:
    async def _go():
        db = await _db()
        upload_dir = Path('/app/backend/uploads/finance')
        upload_dir.mkdir(parents=True, exist_ok=True)
        import_id = str(uuid.uuid4())
        fp = upload_dir / f'{import_id}.xlsx'
        fp.write_bytes(content)
        await db.finance_imports.insert_one({
            'id': import_id,
            'type': 'overdue_balances',
            'source_method': 'manual_upload',
            'filename': filename,
            'file_hash': uuid.uuid4().hex,
            'as_of_date': datetime.now(timezone.utc).date().isoformat(),
            'uploaded_by': 'test',
            'uploaded_at': datetime.now(timezone.utc).isoformat(),
            'status': 'pending_approval',
            'original_file_path': str(fp),
            'totals': {'clients': 3, 'documents': doc_count_hint,
                       'total_balance': 0, 'total_overdue': 0},
            'warnings': ['seed'],
            'errors': [],
            'approved_by': None,
            'approved_at': None,
        })
        return import_id
    return asyncio.run(_go())


def _seed_docs(codes):
    async def _go():
        db = await _db()
        now = datetime.now(timezone.utc).isoformat()
        docs = [{
            'id': f'{c}_PREV{i:02d}',
            'client_id': None,
            'genes_code': c,
            'document_type': 'FT',
            'document_number': f'PREV{i:02d}',
            'invoice_date': '2025-11-01',
            'due_date': '2025-12-01',
            'amount_original': 100.0,
            'amount_open': 100.0,
            'amount_overdue': 100.0,
            'days_overdue': 30,
            'classification': 'collectable',
            'effective_classification': 'collectable',
            'manually_marked_collectable': False,
            'manual_action': None,
            'last_import_id': 'seed',
            'created_at': now,
            'updated_at': now,
        } for c in codes for i in range(2)]
        if docs:
            await db.finance_documents.insert_many(docs)
    asyncio.run(_go())


@pytest.fixture(scope='module', autouse=True)
def _isolate():
    async def snap():
        db = await _db()
        return await db.finance_documents.find({}, {'_id': 0}).to_list(20000)
    docs = asyncio.run(snap())
    yield
    async def restore():
        db = await _db()
        await db.finance_documents.delete_many({})
        if docs:
            await db.finance_documents.insert_many([dict(d) for d in docs])
        # Cleanup TSTPREV imports + files
        cur = db.finance_imports.find(
            {'filename': {'$regex': '^TSTPREV-'}}, {'id': 1, 'original_file_path': 1}
        )
        async for imp in cur:
            p = imp.get('original_file_path')
            if p and os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass
        await db.finance_imports.delete_many({'filename': {'$regex': '^TSTPREV-'}})
    asyncio.run(restore())


def _get_preview(import_id):
    return requests.get(
        f'{BASE_URL}/api/finance/imports/{import_id}/preview',
        headers={'Authorization': session.headers['Authorization']}
    )


class TestPreviewValidOverdue:
    def test_valid_overdue_preview(self):
        # Seed some finance_documents so `current` is > 0
        # (we don't clear all — we just add ours)
        _seed_docs([TST_G])
        marker = uuid.uuid4().hex[:8]
        content = _build_overdue_xlsx_valid(marker)
        import_id = _create_pending_import(content, f'TSTPREV-valid-{marker}.xlsx', 1)

        r = _get_preview(import_id)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data['supported'] is True
        assert data['import_type'] == 'overdue_balances'
        assert data['current']['documents'] >= 2, data['current']
        assert data['new']['clients'] == 1, data['new']
        assert data['new']['documents'] == 1, data['new']
        assert data['delta'] is not None
        # Guard não bloqueia (temos 1 doc parsed contra o guard 0 docs)
        assert data['guard_warnings'] == [], data['guard_warnings']


class TestPreviewZeroDocsCritical:
    def test_zero_docs_preview_is_critical(self):
        _seed_docs([TST_G])  # garante DB tem docs
        marker = uuid.uuid4().hex[:8]
        content = _build_overdue_xlsx_zero_docs(marker)
        import_id = _create_pending_import(content, f'TSTPREV-zero-{marker}.xlsx', 0)

        r = _get_preview(import_id)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data['supported'] is True
        assert data['new']['documents'] == 0, data['new']
        # Guard warning presente
        assert len(data['guard_warnings']) >= 1, data
        warning_text = ' '.join(data['guard_warnings'])
        assert '0 documentos' in warning_text, warning_text
        # is_critical = True (frontend deve mostrar aviso vermelho)
        assert data['is_critical'] is True, data


class TestPreviewNotFound:
    def test_preview_404_when_import_missing(self):
        r = _get_preview('does-not-exist-xyz')
        assert r.status_code == 404, r.text
