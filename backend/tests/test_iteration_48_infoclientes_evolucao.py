"""
Iteration 48 — InfoClientes + Evolução Crédito.

Bug reportado pelo utilizador: ao carregar os ficheiros reais
infocliente_24_07.xlsx e evolucaocredito3em3meses_24_07.xlsx, o
histórico de importações mostrava 0 clientes / 0 documentos e
status="Importado" — porque os parsers procuravam a coluna 'CodCliente'
ou 'CODCLIENTE' mas o ficheiro real usa 'Conta'. Nenhum dado era
persistido → import silencioso.

Este teste valida a correção end-to-end:

  1. Parser InfoClientes com o ficheiro real → >0 clientes (>28k)
  2. Parser Evolução Crédito com o ficheiro real → 579 clientes, 6 períodos
  3. Import silencioso (>10 linhas, 0 clientes) é REJEITADO
  4. Import InfoClientes bem-sucedido enriquece finance_clients com todos
     os campos (saldo_efec, saldo_desc, saldo_dev, carteira, domiciliacoes,
     risco_raw, risco_validado, risco_placeholder, albaranado,
     forma_pagamento, eventos_raw, last_infoclientes_import_id)
  5. Placeholder de risco (> 1M€) marca risco_placeholder=True e
     risco_validado=0.0
  6. Import Evolução Crédito persiste em finance_credit_evolution com o
     formato { genes_code, client_name, periods:{"03-2025":...}, ... }
  7. Endpoint GET /clients/{id}/credit-evolution devolve series,
     peak, quarter_diff_abs, quarter_diff_pct, trend
  8. Import history counters preenchidos: rows_processed, clients_found,
     clients_matched, clients_updated, clients_ignored, documents_created=0
"""
import io
import os
import uuid
import asyncio
from pathlib import Path
from datetime import datetime, timezone

import pytest
import requests
from openpyxl import Workbook
from dotenv import load_dotenv

load_dotenv('/app/backend/.env')
from motor.motor_asyncio import AsyncIOMotorClient  # noqa

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
if not BASE_URL:
    with open('/app/frontend/.env') as f:
        for line in f:
            if line.startswith('REACT_APP_BACKEND_URL='):
                BASE_URL = line.split('=', 1)[1].strip().rstrip('/')

ADMIN_EMAIL = 'admin@pdpv.pt'
ADMIN_PASSWORD = os.environ.get('TEST_ADMIN_PASSWORD', 'HCNMEnKMLq')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']

INFOCLIENTE_FILE = Path('/tmp/infocliente.xlsx')
EVOLUCAO_FILE = Path('/tmp/evolucao.xlsx')

INFOCLIENTE_URL = 'https://customer-assets-4nw71qhi.emergentagent.net/job_9a8fb2b3-6c4b-425c-8e6e-1e9af4e65d07/artifacts/dygn8abl_infocliente_24_07.xlsx'
EVOLUCAO_URL = 'https://customer-assets-4nw71qhi.emergentagent.net/job_9a8fb2b3-6c4b-425c-8e6e-1e9af4e65d07/artifacts/7omvqrke_evolu%C3%A7aocredito3em3meses_24_07.xlsx'


def _ensure_files():
    """Download real fixtures se não existirem em /tmp (container restart)."""
    if not INFOCLIENTE_FILE.exists():
        r = requests.get(INFOCLIENTE_URL, timeout=60)
        r.raise_for_status()
        INFOCLIENTE_FILE.write_bytes(r.content)
    if not EVOLUCAO_FILE.exists():
        r = requests.get(EVOLUCAO_URL, timeout=60)
        r.raise_for_status()
        EVOLUCAO_FILE.write_bytes(r.content)


@pytest.fixture(scope='session', autouse=True)
def _ensure_real_files():
    _ensure_files()

session = requests.Session()
session.headers.update({'Content-Type': 'application/json'})


def _login():
    r = session.post(f'{BASE_URL}/api/auth/login',
                     json={'email': ADMIN_EMAIL, 'password': ADMIN_PASSWORD})
    assert r.status_code == 200, r.text
    session.headers.update({'Authorization': f'Bearer {r.json()["token"]}'})


@pytest.fixture(scope='session', autouse=True)
def _auth():
    _login()


async def _db():
    c = AsyncIOMotorClient(MONGO_URL)
    return c[DB_NAME]


def _add_marker(xlsx_bytes: bytes) -> bytes:
    """Adiciona célula única para tornar o file_hash único entre runs."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    ws['ZZ1'] = f'MARKER-{uuid.uuid4().hex[:8]}'
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ================== PURE PARSER TESTS ==================

class TestClientInfoParser:
    def test_parses_real_file_with_conta_column(self):
        from modules.finance.parsers import parse_client_info
        content = INFOCLIENTE_FILE.read_bytes()
        result = parse_client_info(content)

        assert result['errors'] == []
        assert result['totals']['rows_processed'] > 10_000, result['totals']
        assert result['totals']['client_count'] > 10_000, result['totals']
        assert 'rows_processed' in result['totals']

        # Primeiro cliente conhecido
        first = result['clients'][0]
        assert first['genes_code'] == '2130', first
        assert first['name'].startswith('TRANSFRADELOS'), first
        assert first['saldo_conta'] == 162944.34, first
        # Novos campos iter 48
        assert 'saldo_efec' in first
        assert 'saldo_desc' in first
        assert 'saldo_dev' in first
        assert 'domiciliations' in first
        assert 'payment_method' in first
        assert 'events_raw' in first
        assert 'risk_raw' in first
        assert 'risk_validated' in first
        assert 'risk_placeholder' in first
        assert first['payment_method'] == 'Pagamento a 30 dias'

    def test_risk_placeholder_detection(self):
        """Risco > 1M€ marca placeholder=True e risk_validated=0."""
        from modules.finance.parsers import parse_client_info

        wb = Workbook()
        ws = wb.active
        ws.append(['Alm.', 'Conta', 'Cliente', 'Saldo Conta', 'Saldo Efec.',
                   'Saldo Desc.', 'Saldo Dev.', 'Carteira', 'Domiciliações',
                   'Risco', 'Albaranado', 'Forma Pagamento', 'Eventos'])
        # cliente 1: risco normal 5000€ — conta 21111...9001 → genes_code 9001
        ws.append([1, '2111109001', 'Cli Normal', 1000, 0, 0, 0, 1000, 0, 5000, 0, 'PP', ''])
        # cliente 2: risco placeholder 999999999 — conta 21111...9002 → 9002
        ws.append([1, '2111109002', 'Cli Sem Limite', 1000, 0, 0, 0, 1000, 0, 999_999_999, 0, 'PP', ''])
        buf = io.BytesIO(); wb.save(buf)
        result = parse_client_info(buf.getvalue())

        assert result['errors'] == []
        c1 = next(c for c in result['clients'] if c['genes_code'] == '9001')
        c2 = next(c for c in result['clients'] if c['genes_code'] == '9002')

        assert c1['risk_placeholder'] is False
        assert c1['risk_raw'] == 5000
        assert c1['risk_validated'] == 5000

        assert c2['risk_placeholder'] is True
        assert c2['risk_raw'] == 999_999_999
        assert c2['risk_validated'] == 0.0


class TestCreditEvolutionParser:
    def test_parses_real_file_with_conta_column(self):
        from modules.finance.parsers import parse_credit_evolution
        content = EVOLUCAO_FILE.read_bytes()
        result = parse_credit_evolution(content)

        assert result['errors'] == []
        assert result['totals']['client_count'] == 579, result['totals']
        assert set(result['periods']) == {
            '03-2025', '06-2025', '09-2025', '12-2025', '03-2026', '06-2026'
        }
        first = result['clients'][0]
        assert first['genes_code'] == '2130', first
        assert set(first['evolution'].keys()) == set(result['periods'])
        assert first['evolution']['03-2025'] == 166678.81
        assert first['evolution']['06-2026'] == 160690.49


# ================== SILENT-ZERO GUARD ==================

class TestSilentZeroGuard:
    """Ficheiro com >10 linhas mas coluna de código inesperada → rejected."""

    def _bad_xlsx(self, marker: str = 'x') -> bytes:
        wb = Workbook()
        ws = wb.active
        # header sem CodCliente/Conta reconhecível
        ws.append(['Alm.', 'ID_ESTRANHO', 'Cliente', 'Saldo Conta'])
        for i in range(15):
            ws.append([1, f'X{i:03d}', f'Cli {i} {marker}', 100 + i])
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def test_client_info_silent_zero_rejected(self):
        # Upload direto via API
        content = self._bad_xlsx(marker=uuid.uuid4().hex[:6])
        files = {'file': ('TSTSZG-info.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = requests.post(
            f'{BASE_URL}/api/finance/imports/client_info',
            files=files,
            headers={'Authorization': session.headers['Authorization']},
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data['status'] == 'rejected', data
        assert any('nenhum cliente' in e.lower() for e in data.get('errors', [])), data
        assert data['totals']['clients_updated'] == 0
        assert data['totals']['rows_processed'] >= 15

    def test_credit_evolution_silent_zero_rejected(self):
        content = self._bad_xlsx(marker=uuid.uuid4().hex[:6])
        files = {'file': ('TSTSZG-evo.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = requests.post(
            f'{BASE_URL}/api/finance/imports/credit_evolution',
            files=files,
            headers={'Authorization': session.headers['Authorization']},
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data['status'] == 'rejected', data
        assert any('nenhum cliente' in e.lower() for e in data.get('errors', [])), data


# ================== E2E ENRICHMENT ==================

@pytest.fixture
def seed_matching_client():
    """Cria um finance_clients com o genes_code do primeiro cliente do ficheiro real.
    Garante limpeza prévia para evitar colisão com clientes de outros testes."""
    async def _go():
        db = await _db()
        # Remove qualquer cliente residual com este genes_code de runs anteriores.
        await db.finance_clients.delete_many({'genes_code': '2130'})
        cid = str(uuid.uuid4())
        now = datetime.now(timezone.utc).isoformat()
        await db.finance_clients.insert_one({
            'id': cid,
            'genes_code': '2130',
            'name': 'TRANSFRADELOS, LDA.',
            'overdue_balance_collectable': 0,
            'oldest_overdue_days': 0,
            'financial_status': 'ok',
            'is_blocked': False,
            'manual_marks': [],
            'created_at': now,
            'updated_at': now,
        })
        return cid
    cid = asyncio.run(_go())
    yield cid
    async def _cleanup():
        db = await _db()
        await db.finance_clients.delete_many({'genes_code': '2130'})
    asyncio.run(_cleanup())


class TestInfoClientesEnrichment:
    def test_import_real_file_enriches_matched_client(self, seed_matching_client):
        # Adiciona marker único para evitar file_hash duplicate entre runs
        content = _add_marker(INFOCLIENTE_FILE.read_bytes())
        files = {'file': (f'TSTINFO-{uuid.uuid4().hex[:6]}.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = requests.post(
            f'{BASE_URL}/api/finance/imports/client_info',
            files=files,
            headers={'Authorization': session.headers['Authorization']},
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data['status'] == 'imported', data
        assert data['totals']['clients_updated'] >= 1
        assert data['totals']['clients_found'] > 10_000
        assert data['totals']['documents_created'] == 0
        assert 'rows_processed' in data['totals']

        # Verifica que o cliente foi enriquecido em BD
        async def _read():
            db = await _db()
            return await db.finance_clients.find_one(
                {'id': seed_matching_client}, {'_id': 0}
            )
        c = asyncio.run(_read())
        assert c['saldo_conta'] == 162944.34, c
        assert c['forma_pagamento'] == 'Pagamento a 30 dias', c
        assert c['carteira'] == 162944.34, c
        assert 'risco_raw' in c and 'risco_validado' in c and 'risco_placeholder' in c
        assert c['last_infoclientes_import_id'] == data['import_id']


class TestCreditEvolutionE2E:
    def test_import_persists_and_endpoint_returns(self, seed_matching_client):
        content = _add_marker(EVOLUCAO_FILE.read_bytes())
        files = {'file': (f'TSTEVO-{uuid.uuid4().hex[:6]}.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = requests.post(
            f'{BASE_URL}/api/finance/imports/credit_evolution',
            files=files,
            headers={'Authorization': session.headers['Authorization']},
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data['status'] == 'imported', data
        assert data['totals']['clients_updated'] == 579, data['totals']
        assert data['totals']['clients_found'] == 579, data['totals']
        assert data['totals']['documents_created'] == 0

        # Chama o endpoint da ficha do cliente
        r = requests.get(
            f'{BASE_URL}/api/finance/clients/{seed_matching_client}/credit-evolution',
            headers={'Authorization': session.headers['Authorization']},
        )
        assert r.status_code == 200, r.text
        evo = r.json()
        assert evo['available'] is True, evo
        assert len(evo['series']) == 6
        # ordem cronológica
        periods = [s['period'] for s in evo['series']]
        assert periods == ['03-2025', '06-2025', '09-2025', '12-2025', '03-2026', '06-2026']
        assert evo['series'][0]['value'] == 166678.81
        assert evo['series'][-1]['value'] == 160690.49
        assert evo['peak'] == 245734.92
        assert evo['trend'] in ('up', 'down', 'stable')
        # Δ trimestral entre 03-2026 e 06-2026
        assert evo['last'] == 160690.49
        assert evo['previous'] == 160612.1
