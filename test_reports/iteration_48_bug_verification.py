#!/usr/bin/env python3
"""Focused verification for Iteration 48 finance import zero-count bug."""
import asyncio
import io
import json
import os
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import requests
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook, load_workbook

sys.path.insert(0, "/app/backend")
load_dotenv('/app/backend/.env')

BASE_URL = os.environ.get('BUGTEST_BASE_URL', 'http://localhost:8001').rstrip('/')
MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
INFOCLIENTE_FILE = Path('/tmp/infocliente.xlsx')
EVOLUCAO_FILE = Path('/tmp/evolucao.xlsx')
ADMIN_EMAIL = 'admin@pdpv.pt'
ADMIN_PASSWORD = os.environ.get('TEST_ADMIN_PASSWORD', 'HCNMEnKMLq')
TARGET_GENES = '2111102130'

results = {"checks": [], "errors": []}

def check(name, condition, detail=None):
    entry = {"name": name, "ok": bool(condition), "detail": detail}
    results["checks"].append(entry)
    print(("PASS" if condition else "FAIL") + f" - {name}: {detail}")
    if not condition:
        results["errors"].append(entry)

def add_marker(xlsx_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    ws['ZZ1'] = f'BUGTEST-{uuid.uuid4().hex}'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def bad_xlsx(kind: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    if kind == 'evolution':
        ws.append(['ID_ESTRANHO', 'Cliente', '03-2025', '06-2025'])
        for i in range(15):
            ws.append([f'X{i:03d}', f'Cli {i}', i * 10, i * 20])
    else:
        ws.append(['Alm.', 'ID_ESTRANHO', 'Cliente', 'Saldo Conta'])
        for i in range(15):
            ws.append([1, f'X{i:03d}', f'Cli {i}', 100 + i])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

async def snapshot_and_seed():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    backup_clients = await db.finance_clients.find({'genes_code': TARGET_GENES}).to_list(None)
    backup_evo = await db.finance_credit_evolution.find({'genes_code': TARGET_GENES}).to_list(None)
    await db.finance_clients.delete_many({'genes_code': TARGET_GENES})
    await db.finance_credit_evolution.delete_many({'genes_code': TARGET_GENES})
    cid = f'bugtest-{uuid.uuid4()}'
    now = datetime.now(timezone.utc).isoformat()
    await db.finance_clients.insert_one({
        'id': cid,
        'genes_code': TARGET_GENES,
        'name': 'TRANSFRADELOS, LDA. BUGTEST',
        'total_balance': 0.0,
        'overdue_balance_accounting': 0.0,
        'overdue_balance_collectable': 0.0,
        'residual_balance': 0.0,
        'oldest_overdue_days': 0,
        'collection_index': 0.0,
        'financial_status': 'OK',
        'traffic_light': 'GREEN',
        'is_residual_only': False,
        'is_blocked': False,
        'created_at': now,
        'updated_at': now,
    })
    client.close()
    return cid, backup_clients, backup_evo

async def cleanup(cid, backup_clients, backup_evo):
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    await db.finance_clients.delete_many({'genes_code': TARGET_GENES})
    await db.finance_credit_evolution.delete_many({'genes_code': TARGET_GENES})
    if backup_clients:
        for doc in backup_clients:
            await db.finance_clients.insert_one(doc)
    if backup_evo:
        for doc in backup_evo:
            await db.finance_credit_evolution.insert_one(doc)
    client.close()

async def get_client(cid):
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    doc = await db.finance_clients.find_one({'id': cid}, {'_id': 0})
    client.close()
    return doc

async def get_evo_doc():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    doc = await db.finance_credit_evolution.find_one({'genes_code': TARGET_GENES}, {'_id': 0})
    client.close()
    return doc

def main():
    print(f"Using API {BASE_URL}, DB {DB_NAME}")
    # Parser proof with real user files.
    from modules.finance.parsers import parse_client_info, parse_credit_evolution
    info = parse_client_info(INFOCLIENTE_FILE.read_bytes())
    check('parse_client_info real file returns >10000 clients', info['totals'].get('client_count', 0) > 10000, info['totals'])
    check('parse_client_info first client/amount/payment method',
          info['clients'][0]['genes_code'] == TARGET_GENES and info['clients'][0]['saldo_conta'] == 162944.34 and info['clients'][0]['payment_method'] == 'Pagamento a 30 dias',
          info['clients'][0])
    evo = parse_credit_evolution(EVOLUCAO_FILE.read_bytes())
    expected_periods = ['03-2025', '06-2025', '09-2025', '12-2025', '03-2026', '06-2026']
    check('parse_credit_evolution real file returns 579 clients and 6 periods',
          evo['totals'].get('client_count') == 579 and evo['periods'] == expected_periods,
          {"totals": evo['totals'], "periods": evo['periods']})

    s = requests.Session()
    login = s.post(f'{BASE_URL}/api/auth/login', json={'email': ADMIN_EMAIL, 'password': ADMIN_PASSWORD}, timeout=30)
    check('admin login works', login.status_code == 200, login.text[:200])
    login.raise_for_status()
    s.headers.update({'Authorization': f"Bearer {login.json()['token']}"})

    cid, backup_clients, backup_evo = asyncio.run(snapshot_and_seed())
    try:
        # Silent-zero guard first.
        for import_type, kind in [('client_info', 'info'), ('credit_evolution', 'evolution')]:
            r = s.post(
                f'{BASE_URL}/api/finance/imports/{import_type}',
                files={'file': (f'bugtest-bad-{import_type}-{uuid.uuid4().hex}.xlsx', bad_xlsx(kind), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=60,
            )
            data = r.json() if r.headers.get('content-type', '').startswith('application/json') else {"text": r.text}
            check(f'silent-zero guard rejects {import_type} with HTTP 200',
                  r.status_code == 200 and data.get('status') == 'rejected' and any('Nenhum cliente encontrado' in e for e in data.get('errors', [])),
                  data)

        # Successful InfoClientes import.
        r = s.post(
            f'{BASE_URL}/api/finance/imports/client_info',
            files={'file': (f'bugtest-infocliente-{uuid.uuid4().hex}.xlsx', add_marker(INFOCLIENTE_FILE.read_bytes()), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            timeout=180,
        )
        info_upload = r.json() if r.headers.get('content-type', '').startswith('application/json') else {"text": r.text}
        check('POST client_info real file imported with non-zero clients and counters',
              r.status_code == 200 and info_upload.get('status') == 'imported' and info_upload.get('totals', {}).get('clients_found', 0) > 10000 and info_upload.get('totals', {}).get('clients_updated', 0) >= 1 and info_upload.get('totals', {}).get('documents_created') == 0,
              info_upload)
        cdoc = asyncio.run(get_client(cid))
        check('InfoClientes enriched finance_clients for genes_code 2111102130',
              cdoc and cdoc.get('saldo_conta') == 162944.34 and cdoc.get('forma_pagamento') == 'Pagamento a 30 dias' and cdoc.get('carteira') == 162944.34 and all(k in cdoc for k in ['risco_raw', 'risco_validado', 'risco_placeholder', 'last_infoclientes_import_id']),
              cdoc)

        # Successful credit evolution import and endpoint.
        r = s.post(
            f'{BASE_URL}/api/finance/imports/credit_evolution',
            files={'file': (f'bugtest-evolucao-{uuid.uuid4().hex}.xlsx', add_marker(EVOLUCAO_FILE.read_bytes()), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
            timeout=180,
        )
        evo_upload = r.json() if r.headers.get('content-type', '').startswith('application/json') else {"text": r.text}
        check('POST credit_evolution real file imported with 579 clients and docs=0',
              r.status_code == 200 and evo_upload.get('status') == 'imported' and evo_upload.get('totals', {}).get('clients_updated') == 579 and evo_upload.get('totals', {}).get('documents_created') == 0,
              evo_upload)
        evodoc = asyncio.run(get_evo_doc())
        check('finance_credit_evolution persisted periods for target client',
              evodoc and evodoc.get('periods', {}).get('03-2025') == 166678.81 and evodoc.get('periods', {}).get('06-2026') == 160690.49 and evodoc.get('last_import_id') == evo_upload.get('import_id'),
              evodoc)
        r = s.get(f'{BASE_URL}/api/finance/clients/{cid}/credit-evolution', timeout=60)
        endpoint = r.json() if r.headers.get('content-type', '').startswith('application/json') else {"text": r.text}
        check('GET client credit-evolution endpoint returns ordered 6-point series and metrics',
              r.status_code == 200 and endpoint.get('available') is True and [p['period'] for p in endpoint.get('series', [])] == expected_periods and endpoint.get('peak') == 245734.92 and endpoint.get('last') == 160690.49 and endpoint.get('previous') == 160612.1,
              endpoint)

        # Import history response that powers /finance/imports UI.
        ids = [info_upload.get('import_id'), evo_upload.get('import_id')]
        r = s.get(f'{BASE_URL}/api/finance/imports?limit=20&offset=0', timeout=60)
        imports_data = r.json()
        rows = [i for i in imports_data.get('imports', []) if i.get('id') in ids]
        check('/api/finance/imports list includes newly imported rows', len(rows) == 2, rows)
        # This is the exact API contract used by frontend cells. Extra counters should survive.
        counters_visible = all(('clients_updated' in row.get('totals', {}) and 'documents_created' in row.get('totals', {})) for row in rows)
        check('/api/finance/imports preserves clients_updated/documents_created for frontend cells', counters_visible, rows)
    finally:
        asyncio.run(cleanup(cid, backup_clients, backup_evo))

    results['success'] = not results['errors']
    print('RESULT_JSON=' + json.dumps(results, ensure_ascii=False, default=str))
    return 0 if results['success'] else 1

if __name__ == '__main__':
    raise SystemExit(main())
