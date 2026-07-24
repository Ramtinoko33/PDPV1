import json
import os
import uuid
import asyncio
import io
from datetime import datetime, timezone
from typing import Any, Dict, List

import requests
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook


load_dotenv("/app/backend/.env")


BASE_URL = os.environ.get("TEST_BASE_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"
ADMIN_EMAIL = os.environ.get("TEST_ADMIN_EMAIL", "admin@pdpv.pt")
ADMIN_PASSWORD = os.environ.get("TEST_ADMIN_PASSWORD", "HCNMEnKMLq")
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "test_database")
INFOCLIENTE_FILE = "/tmp/infocliente.xlsx"
EVOLUCAO_FILE = "/tmp/evolucao.xlsx"


def assert_true(condition: bool, message: str):
    if not condition:
        raise AssertionError(message)


def _add_marker(path: str) -> bytes:
    """Make uploads unique so duplicate file_hash does not block repeated retests."""
    wb = load_workbook(path)
    ws = wb.active
    ws["ZZ1"] = f"QA-MARKER-{uuid.uuid4().hex[:10]}"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


async def _seed_matching_client() -> str:
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    existing = await db.finance_clients.find_one({"genes_code": "2111102130"}, {"_id": 0})
    if existing:
        client.close()
        return existing["id"]
    cid = f"qa-ret48-{uuid.uuid4()}"
    now = datetime.now(timezone.utc).isoformat()
    await db.finance_clients.insert_one({
        "id": cid,
        "genes_code": "2111102130",
        "name": "TRANSFRADELOS, LDA. QA RETEST",
        "total_balance": 0.0,
        "overdue_balance_accounting": 0.0,
        "overdue_balance_collectable": 0.0,
        "residual_balance": 0.0,
        "oldest_overdue_days": 0,
        "collection_index": 0.0,
        "financial_status": "OK",
        "traffic_light": "GREEN",
        "is_residual_only": False,
        "is_blocked": False,
        "created_at": now,
        "updated_at": now,
    })
    client.close()
    return cid


def main() -> Dict[str, Any]:
    s = requests.Session()
    login = s.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=20)
    assert_true(login.status_code == 200, f"login failed: {login.status_code} {login.text[:300]}")
    s.headers.update({"Authorization": f"Bearer {login.json()['token']}"})

    seeded_client_id = asyncio.run(_seed_matching_client())

    info_upload = s.post(
        f"{API}/finance/imports/client_info",
        files={"file": (f"QA-INFO-{uuid.uuid4().hex[:6]}.xlsx", _add_marker(INFOCLIENTE_FILE), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=120,
    )
    assert_true(info_upload.status_code == 200, f"client_info upload failed: {info_upload.status_code} {info_upload.text[:500]}")
    info_data = info_upload.json()
    assert_true(info_data.get("status") == "imported", f"client_info upload not imported: {info_data}")
    assert_true(info_data.get("totals", {}).get("clients_found") == 28620, f"client_info real file did not parse 28620 clients: {info_data}")
    assert_true(info_data.get("totals", {}).get("clients_updated", 0) >= 1, f"client_info did not update seeded client: {info_data}")

    evo_upload = s.post(
        f"{API}/finance/imports/credit_evolution",
        files={"file": (f"QA-EVO-{uuid.uuid4().hex[:6]}.xlsx", _add_marker(EVOLUCAO_FILE), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=120,
    )
    assert_true(evo_upload.status_code == 200, f"credit_evolution upload failed: {evo_upload.status_code} {evo_upload.text[:500]}")
    evo_data = evo_upload.json()
    assert_true(evo_data.get("status") == "imported", f"credit_evolution upload not imported: {evo_data}")
    assert_true(evo_data.get("totals", {}).get("clients_found") == 579, f"credit_evolution real file did not parse 579 clients: {evo_data}")
    assert_true(evo_data.get("totals", {}).get("periods") == 6, f"credit_evolution did not parse 6 periods: {evo_data}")

    imports_res = s.get(f"{API}/finance/imports?limit=100&offset=0", timeout=30)
    assert_true(imports_res.status_code == 200, f"imports failed: {imports_res.status_code} {imports_res.text[:500]}")
    imports_payload = imports_res.json()
    imports: List[Dict[str, Any]] = imports_payload.get("imports", [])

    required_common = {
        "rows_processed",
        "clients_found",
        "clients_matched",
        "clients_updated",
        "clients_ignored",
        "documents_created",
    }
    client_info = next((i for i in imports if i.get("type") == "client_info" and i.get("totals")), None)
    credit_evo = next((i for i in imports if i.get("type") == "credit_evolution" and i.get("totals")), None)
    assert_true(client_info is not None, "no client_info import found in /finance/imports")
    assert_true(credit_evo is not None, "no credit_evolution import found in /finance/imports")

    ci_totals = client_info["totals"]
    ce_totals = credit_evo["totals"]
    assert_true(required_common.issubset(ci_totals.keys()), f"client_info totals missing keys: {required_common - set(ci_totals.keys())}")
    assert_true(required_common.union({"periods"}).issubset(ce_totals.keys()), f"credit_evolution totals missing keys: {required_common.union({'periods'}) - set(ce_totals.keys())}")
    assert_true(ci_totals.get("clients_found", 0) >= 28620, f"client_info clients_found too low: {ci_totals}")
    assert_true(ci_totals.get("documents_created") == 0, f"client_info documents_created not 0: {ci_totals}")
    assert_true(ce_totals.get("clients_found") == 579, f"credit_evolution clients_found not 579: {ce_totals}")
    assert_true(ce_totals.get("periods") == 6, f"credit_evolution periods not 6: {ce_totals}")
    assert_true(ce_totals.get("documents_created") == 0, f"credit_evolution documents_created not 0: {ce_totals}")

    evo_res = s.get(f"{API}/finance/clients/{seeded_client_id}/credit-evolution", timeout=30)
    assert_true(evo_res.status_code == 200, f"credit evolution endpoint failed: {evo_res.status_code} {evo_res.text[:500]}")
    evo = evo_res.json()
    assert_true(evo.get("available") is True, f"credit evolution not available: {evo}")
    periods = [p.get("period") for p in evo.get("series", [])]
    assert_true(periods == ["03-2025", "06-2025", "09-2025", "12-2025", "03-2026", "06-2026"], f"period order wrong: {periods}")
    for key in ["peak", "last", "previous", "quarter_diff_abs", "quarter_diff_pct", "trend"]:
        assert_true(key in evo, f"credit evolution missing {key}: {evo}")
    assert_true(evo.get("last") == 160690.49, f"unexpected last credit value: {evo}")
    assert_true(evo.get("previous") == 160612.1, f"unexpected previous credit value: {evo}")
    assert_true(evo.get("peak") == 245734.92, f"unexpected peak credit value: {evo}")

    return {
        "base_url": BASE_URL,
        "imports_checked": len(imports),
        "client_info_import": {"id": client_info["id"], "totals": ci_totals},
        "credit_evolution_import": {"id": credit_evo["id"], "totals": ce_totals},
        "uploaded_client_info": {"import_id": info_data.get("import_id"), "totals": info_data.get("totals")},
        "uploaded_credit_evolution": {"import_id": evo_data.get("import_id"), "totals": evo_data.get("totals")},
        "client_id": seeded_client_id,
        "credit_evolution": evo,
    }


if __name__ == "__main__":
    result = main()
    print(json.dumps(result, indent=2, ensure_ascii=False))