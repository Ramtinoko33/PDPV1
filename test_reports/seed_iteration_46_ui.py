"""Seed pending imports + files for iteration 46 frontend testing."""
import os, io, uuid, asyncio
from datetime import datetime, timezone
from pathlib import Path
from dotenv import load_dotenv
load_dotenv('/app/backend/.env')
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']


def _xlsx_overdue_zero_docs(marker):
    wb = Workbook(); ws = wb.active
    ws.append(['Cliente','Cód. Cliente','Localidade','Região','Email','Telefone1','Telefone2','Importe Total Vencido','Saldo Cliente'])
    for i in range(3):
        ws.append([f'Cli{i}', f'ZERO{i}', 'Lx','Sul',None,None,None,0.0,0.0])
    ws['Z1']=marker
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


def _xlsx_overdue_valid(marker):
    wb = Workbook(); ws = wb.active
    ws.append(['Cliente','Cód. Cliente','Localidade','Região','Email','Telefone1','Telefone2','Importe Total Vencido','Saldo Cliente'])
    ws.append([f'Cli {marker}','UITST1','Lx','Sul','x@x.pt',None,None,250.0,250.0])
    ws.append(['','Documento','Data da fatura','Data Vencimento','CódSede','Sede','Dias Vencidos','Importe Vencimiento','Vencido Factura'])
    ws.append(['','999/ui1', datetime(2025,11,1), datetime(2025,12,1),'01','Sede',30,250.0,250.0])
    ws['Z1']=marker
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


async def main():
    c = AsyncIOMotorClient(MONGO_URL)
    db = c[DB_NAME]

    # Cleanup previous UI seeds
    cur = db.finance_imports.find({'filename': {'$regex':'^UITST46-'}}, {'id':1,'original_file_path':1})
    async for imp in cur:
        p = imp.get('original_file_path')
        if p and os.path.exists(p):
            try: os.remove(p)
            except Exception: pass
    await db.finance_imports.delete_many({'filename': {'$regex':'^UITST46-'}})

    upload_dir = Path('/app/backend/uploads/finance'); upload_dir.mkdir(parents=True, exist_ok=True)

    created = []

    # 1) Pending overdue with 0 docs → guard_warnings + is_critical
    for label, xlsx_bytes, doc_count in [
        ('zero', _xlsx_overdue_zero_docs('m1'), 0),
        ('valid', _xlsx_overdue_valid('m2'), 1),
    ]:
        import_id = str(uuid.uuid4())
        fp = upload_dir / f'{import_id}.xlsx'
        fp.write_bytes(xlsx_bytes)
        doc = {
            'id': import_id,
            'type': 'overdue_balances',
            'source_method': 'manual_upload',
            'filename': f'UITST46-{label}-{import_id[:6]}.xlsx',
            'file_hash': uuid.uuid4().hex,
            'as_of_date': datetime.now(timezone.utc).date().isoformat(),
            'uploaded_by': 'admin@pdpv.pt',
            'uploaded_by_name': 'Admin Test',
            'uploaded_at': datetime.now(timezone.utc).isoformat(),
            'status': 'pending_approval',
            'original_file_path': str(fp),
            'totals': {'clients': 3, 'documents': doc_count, 'total_balance': 0, 'total_overdue': 0},
            'warnings': ['seed'],
            'errors': [],
            'approved_by': None, 'approved_at': None,
        }
        await db.finance_imports.insert_one(doc)
        created.append({'label': label, 'import_id': import_id, 'filename': doc['filename']})
        print(f"SEEDED {label}: {import_id} filename={doc['filename']}")

    print("DONE.")

asyncio.run(main())
