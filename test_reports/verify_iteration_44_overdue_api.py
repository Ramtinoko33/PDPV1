"""
Focused backend/API verification for Iteration 44 overdue_balances safety guard.

This script uses the preview API (REACT_APP_BACKEND_URL) plus MongoDB state checks.
It creates only TSTOVAPI-* test data and restores finance_documents/finance_clients
snapshots at the end to protect the preview dataset.
"""
import asyncio
import io
import json
import os
import uuid
from datetime import datetime, timezone, date
from pathlib import Path

import requests
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook


ROOT = Path("/app")
load_dotenv(ROOT / "backend" / ".env")


def read_frontend_api_url() -> str:
    env_path = ROOT / "frontend" / ".env"
    for line in env_path.read_text().splitlines():
        if line.startswith("REACT_APP_BACKEND_URL="):
            return line.split("=", 1)[1].strip().rstrip("/")
    raise RuntimeError("REACT_APP_BACKEND_URL not found")


BASE_URL = read_frontend_api_url()
MONGO_URL = os.environ["MONGO_URL"].strip('"')
DB_NAME = os.environ["DB_NAME"].strip('"')
ADMIN_EMAIL = "admin@pdpv.pt"
ADMIN_PASSWORD = os.environ.get("TEST_ADMIN_PASSWORD", "HCNMEnKMLq")
TODAY = date.today().isoformat()
PREFIX = "TSTOVAPI"


def build_overdue_xlsx_with_docs(marker: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Cliente", "Cód. Cliente", "Localidade", "Região", "Email",
        "Telefone1", "Telefone2", "Importe Total Vencido", "Saldo Cliente",
    ])
    ws.append([f"Cliente Teste {marker}", f"{PREFIX}V", "Lisboa", "Sul", "test@example.pt", "", "", 123.45, 123.45])
    ws.append(["", "Documento", "Data da fatura", "Data Vencimento", "CódSede", "Sede", "Dias Vencidos", "Importe Vencimiento", "Vencido Factura"])
    ws.append(["", f"026/{marker}", datetime(2025, 11, 1), datetime(2025, 12, 1), "01", "Sede", 30, 123.45, 123.45])
    ws["Z1"] = marker
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_overdue_xlsx_zero_docs(marker: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Cliente", "Cód. Cliente", "Localidade", "Região", "Email",
        "Telefone1", "Telefone2", "Importe Total Vencido", "Saldo Cliente",
    ])
    for i in range(5):
        ws.append([f"Cliente Zero {i} {marker}", f"{PREFIX}Z{i:02d}", "Lisboa", "Sul", "", "", "", 0.0, 0.0])
    ws["Z1"] = marker
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def cleanup_test_data(db, restore_docs=None, restore_clients=None, restore_health_marker=Ellipsis):
    imports = await db.finance_imports.find({"filename": {"$regex": f"^{PREFIX}-"}}, {"_id": 0, "id": 1, "original_file_path": 1}).to_list(100)
    import_ids = [i["id"] for i in imports]
    for imp in imports:
        p = imp.get("original_file_path")
        if p and Path(p).exists():
            try:
                Path(p).unlink()
            except Exception:
                pass
    if import_ids:
        await db.finance_client_daily_metrics.delete_many({"import_id": {"$in": import_ids}})
    await db.finance_imports.delete_many({"filename": {"$regex": f"^{PREFIX}-"}})
    await db.finance_documents.delete_many({"genes_code": {"$regex": f"^{PREFIX}"}})
    await db.finance_clients.delete_many({"genes_code": {"$regex": f"^{PREFIX}"}})
    if restore_docs is not None:
        await db.finance_documents.delete_many({})
        if restore_docs:
            await db.finance_documents.insert_many([dict(d) for d in restore_docs])
    if restore_clients is not None:
        await db.finance_clients.delete_many({})
        if restore_clients:
            await db.finance_clients.insert_many([dict(c) for c in restore_clients])
    if restore_health_marker is not Ellipsis:
        if restore_health_marker is None:
            await db.finance_data_health.delete_one({"source_type": "overdue_balances"})
        else:
            await db.finance_data_health.update_one(
                {"source_type": "overdue_balances"},
                {"$set": restore_health_marker},
                upsert=True,
            )


async def seed_existing_docs(db):
    now = datetime.now(timezone.utc).isoformat()
    docs = []
    for code in [f"{PREFIX}A", f"{PREFIX}B"]:
        for i in range(3):
            docs.append({
                "id": f"{code}_SEED{i}",
                "client_id": None,
                "genes_code": code,
                "document_type": "FT",
                "document_number": f"SEED{i}",
                "invoice_date": "2025-11-01",
                "due_date": "2025-12-01",
                "amount_original": 100.0,
                "amount_open": 100.0,
                "amount_overdue": 100.0,
                "days_overdue": 30,
                "classification": "collectable",
                "effective_classification": "collectable",
                "manually_marked_collectable": False,
                "manual_action": None,
                "last_import_id": f"{PREFIX}-seed",
                "created_at": now,
                "updated_at": now,
            })
    await db.finance_documents.insert_many(docs)


def login() -> str:
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=30)
    r.raise_for_status()
    return r.json()["token"]


def upload_overdue(token: str, content: bytes, filename: str):
    return requests.post(
        f"{BASE_URL}/api/finance/imports/overdue_balances",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=60,
    )


def approve(token: str, import_id: str):
    return requests.post(
        f"{BASE_URL}/api/finance/imports/{import_id}/approve",
        headers={"Authorization": f"Bearer {token}"},
        timeout=60,
    )


async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    evidence = {"base_url": BASE_URL, "checks": []}
    docs_snapshot = await db.finance_documents.find({}, {"_id": 0}).to_list(50000)
    clients_snapshot = await db.finance_clients.find({}, {"_id": 0}).to_list(50000)
    health_snapshot = await db.finance_data_health.find_one({"source_type": "overdue_balances"}, {"_id": 0})

    try:
        await cleanup_test_data(db)
        token = login()
        evidence["checks"].append({"name": "login", "ok": True})

        # 1) Direct upload of 0-doc overdue file must be rejected and not mutate docs/clients.
        await seed_existing_docs(db)
        docs_before = await db.finance_documents.count_documents({})
        clients_before = await db.finance_clients.count_documents({})
        marker = uuid.uuid4().hex[:8]
        r = upload_overdue(token, build_overdue_xlsx_zero_docs(marker), f"{PREFIX}-zero-{marker}.xlsx")
        data = r.json()
        docs_after = await db.finance_documents.count_documents({})
        clients_after = await db.finance_clients.count_documents({})
        evidence["checks"].append({
            "name": "zero_docs_upload_guard",
            "http_status": r.status_code,
            "response_status": data.get("status"),
            "errors": data.get("errors"),
            "totals": data.get("totals"),
            "docs_before": docs_before,
            "docs_after": docs_after,
            "clients_before": clients_before,
            "clients_after": clients_after,
            "ok": r.status_code == 200 and data.get("status") == "rejected" and "0 documentos" in " ".join(data.get("errors", [])) and docs_before == docs_after and clients_before == clients_after,
        })

        # 2) Approving an already pending 0-doc import must return 400 with readable detail, not 500.
        docs_before = await db.finance_documents.count_documents({})
        clients_before = await db.finance_clients.count_documents({})
        marker = uuid.uuid4().hex[:8]
        import_id = str(uuid.uuid4())
        upload_dir = ROOT / "backend" / "uploads" / "finance_imports"
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / f"{PREFIX}-pending-{import_id}.xlsx"
        file_path.write_bytes(build_overdue_xlsx_zero_docs(marker))
        await db.finance_imports.insert_one({
            "id": import_id,
            "type": "overdue_balances",
            "source_method": "manual_upload",
            "filename": f"{PREFIX}-pending-{marker}.xlsx",
            "file_hash": uuid.uuid4().hex,
            "as_of_date": TODAY,
            "uploaded_by": "test-verification",
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "status": "pending_approval",
            "original_file_path": str(file_path),
            "totals": {"clients": 5, "documents": 0, "total_balance": 0, "total_overdue": 0},
            "warnings": ["seed pending import for guard verification"],
            "errors": [],
            "approved_by": None,
            "approved_at": None,
        })
        r = approve(token, import_id)
        body = r.json()
        imp_after = await db.finance_imports.find_one({"id": import_id}, {"_id": 0, "status": 1, "errors": 1})
        docs_after = await db.finance_documents.count_documents({})
        clients_after = await db.finance_clients.count_documents({})
        detail = body.get("detail", "") if isinstance(body, dict) else ""
        evidence["checks"].append({
            "name": "approve_pending_zero_docs_returns_400",
            "http_status": r.status_code,
            "detail": detail,
            "import_after": imp_after,
            "docs_before": docs_before,
            "docs_after": docs_after,
            "clients_before": clients_before,
            "clients_after": clients_after,
            "ok": r.status_code == 400 and "0 documentos" in detail and docs_before == docs_after and clients_before == clients_after and imp_after.get("status") == "rejected",
        })

        # 3) Valid overdue file with >0 parsed documents must not be rejected by the new guard.
        marker = uuid.uuid4().hex[:8]
        r = upload_overdue(token, build_overdue_xlsx_with_docs(marker), f"{PREFIX}-valid-{marker}.xlsx")
        data = r.json()
        evidence["checks"].append({
            "name": "valid_overdue_import_not_rejected",
            "http_status": r.status_code,
            "response_status": data.get("status"),
            "success": data.get("success"),
            "totals": data.get("totals"),
            "errors": data.get("errors"),
            "ok": r.status_code == 200 and data.get("status") != "rejected" and (data.get("totals") or {}).get("documents", 0) >= 1 and "0 documentos" not in " ".join(data.get("errors", [])),
        })

        evidence["ok"] = all(c.get("ok") for c in evidence["checks"])
        out_path = ROOT / "test_reports" / "iteration_44_api_evidence.json"
        out_path.write_text(json.dumps(evidence, indent=2, ensure_ascii=False))
        print(json.dumps(evidence, indent=2, ensure_ascii=False))
        if not evidence["ok"]:
            raise SystemExit(1)
    finally:
        await cleanup_test_data(db, restore_docs=docs_snapshot, restore_clients=clients_snapshot, restore_health_marker=health_snapshot)
        client.close()


if __name__ == "__main__":
    asyncio.run(main())